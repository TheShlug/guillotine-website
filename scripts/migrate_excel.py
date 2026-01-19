"""
Excel to JSON Migration Script for Guillotine League Website
Extracts historical data from Excel file for 2023 and 2024 seasons.
"""

import openpyxl
import json
from pathlib import Path
import statistics

EXCEL_PATH = Path(__file__).parent.parent.parent / "Gheny+ Guillotine League 2025.xlsx"
OUTPUT_DIR = Path(__file__).parent.parent / "frontend" / "data"


def extract_2024_data(wb):
    """Extract 2024 season data from Excel."""
    sheet = wb['2024 Guillotine']

    managers = []

    # Data is in columns 25-49 (Y-AW)
    # Row 1 is header, rows 2-19 are managers, rows 20-25 are summary stats
    for row in range(2, 20):  # 18 managers
        username = sheet.cell(row=row, column=25).value
        if not username or username in ['High score', '75th percentile', 'Median score', '25th percentile', 'CHOP Score', 'CHOP differenti']:
            continue

        # Weekly scores from columns 26-42 (wk1-wk17)
        weekly_scores = {}
        for wk in range(1, 18):
            score = sheet.cell(row=row, column=25 + wk).value
            weekly_scores[str(wk)] = float(score) if score is not None else None

        # Chop week and draft position from columns 48-49
        chop_week = sheet.cell(row=row, column=48).value
        draft_pos = sheet.cell(row=row, column=49).value

        # chop_week of 18 means winner (survived all 17 weeks) - set to None
        if chop_week == 18:
            chop_week = None

        # FAAB - need to find it from left side columns (column 19)
        # Match by username from column 1
        faab = None
        for search_row in range(2, 20):
            left_username = sheet.cell(row=search_row, column=1).value
            if left_username == username or (left_username and username and left_username.lower() == username.lower()):
                faab = sheet.cell(row=search_row, column=19).value
                break

        # If not found by exact match, search column 25 on left side
        if faab is None:
            for search_row in range(2, 20):
                left_username = sheet.cell(row=search_row, column=25).value
                if left_username == username:
                    # Look for corresponding FAAB in original data
                    # The left side has different usernames in column 1
                    pass

        managers.append({
            "user_name": username,
            "draft_position": int(draft_pos) if draft_pos else None,
            "chop_week": int(chop_week) if chop_week else None,
            "faab_remaining": int(faab) if faab is not None else 0,
            "weekly_scores": weekly_scores
        })

    # Extract weekly stats from rows 20-25
    weekly_stats = {}
    stat_rows = {
        20: "high_score",
        21: "percentile_75",
        22: "median",
        23: "percentile_25",
        24: "chop_score",
        25: "chop_differential"
    }

    for wk in range(1, 18):
        col = 25 + wk
        stats = {}
        for row_num, stat_name in stat_rows.items():
            val = sheet.cell(row=row_num, column=col).value
            stats[stat_name] = float(val) if val is not None else None
        weekly_stats[str(wk)] = stats

    # Calculate avg_above_chop for each manager
    for manager in managers:
        chop_week = manager["chop_week"]
        end_week = chop_week if chop_week else 17

        diffs = []
        for wk in range(1, end_week + 1):
            score = manager["weekly_scores"].get(str(wk))
            chop = weekly_stats.get(str(wk), {}).get("chop_score")
            if score is not None and chop is not None:
                diffs.append(score - chop)

        manager["avg_above_chop"] = round(statistics.mean(diffs), 1) if diffs else 0

    # Sort: eliminated first (by chop_week), then survivors (by avg_above_chop ascending)
    managers.sort(key=lambda m: (
        0 if m["chop_week"] else 1,
        m["chop_week"] or 999,
        m["avg_above_chop"]
    ))

    return {
        "season": 2024,
        "champion": "TheShlug",
        "starting_faab": 1000,
        "managers": managers,
        "weekly_stats": weekly_stats
    }


def extract_2023_data(wb):
    """Extract 2023 season data from Excel."""
    sheet = wb['2023 Guillotine']

    managers = []

    # 2023 structure: Column 1 = chop week, Column 2 = username, Columns 3-19 = wk1-wk17 scores (above chop)
    # Columns 25-41 have the actual raw scores

    for row in range(2, 20):  # 18 managers
        chop_week = sheet.cell(row=row, column=1).value
        username = sheet.cell(row=row, column=2).value

        if not username or username == 'Avg:':
            continue

        # Weekly scores from columns 25-41 (actual scores, not above-chop)
        weekly_scores = {}
        for wk in range(1, 18):
            score = sheet.cell(row=row, column=24 + wk).value
            weekly_scores[str(wk)] = float(score) if score is not None else None

        # FAAB from column 20
        faab = sheet.cell(row=row, column=20).value

        # chop_week of 18 or no chop_week means winner/survivor
        parsed_chop = int(chop_week) if chop_week and isinstance(chop_week, (int, float)) else None
        if parsed_chop == 18:
            parsed_chop = None

        managers.append({
            "user_name": username,
            "draft_position": None,  # Not available for 2023
            "chop_week": parsed_chop,
            "faab_remaining": int(faab) if faab is not None else 0,
            "weekly_scores": weekly_scores
        })

    # Extract weekly stats from rows 20-25
    weekly_stats = {}
    stat_rows = {
        20: "high_score",
        21: "percentile_75",
        22: "median",
        23: "percentile_25",
        24: "chop_score",
        25: "chop_differential"
    }

    for wk in range(1, 18):
        col = 24 + wk
        stats = {}
        for row_num, stat_name in stat_rows.items():
            val = sheet.cell(row=row_num, column=col).value
            stats[stat_name] = float(val) if val is not None else None
        weekly_stats[str(wk)] = stats

    # Calculate avg_above_chop for each manager
    for manager in managers:
        chop_week = manager["chop_week"]
        end_week = chop_week if chop_week else 17

        diffs = []
        for wk in range(1, end_week + 1):
            score = manager["weekly_scores"].get(str(wk))
            chop = weekly_stats.get(str(wk), {}).get("chop_score")
            if score is not None and chop is not None:
                diffs.append(score - chop)

        manager["avg_above_chop"] = round(statistics.mean(diffs), 1) if diffs else 0

    # Sort: eliminated first (by chop_week), then survivors (by avg_above_chop ascending)
    managers.sort(key=lambda m: (
        0 if m["chop_week"] else 1,
        m["chop_week"] or 999,
        m["avg_above_chop"]
    ))

    return {
        "season": 2023,
        "champion": "TheShlug",
        "starting_faab": 1000,
        "has_draft_positions": False,
        "managers": managers,
        "weekly_stats": weekly_stats
    }


def extract_average_finishes(wb):
    """Extract average finishes data across seasons."""
    sheet = wb['Average finishes']

    managers = []

    for row in range(2, 25):  # Up to ~20 managers
        name = sheet.cell(row=row, column=1).value
        if not name or name in ['Avg:', 'median:']:
            continue

        finish_2023 = sheet.cell(row=row, column=2).value
        finish_2024 = sheet.cell(row=row, column=3).value
        finish_2025 = sheet.cell(row=row, column=4).value
        avg_finish = sheet.cell(row=row, column=5).value

        # Convert '-' to None
        def parse_finish(val):
            if val is None or val == '-':
                return None
            return int(val) if isinstance(val, (int, float)) else None

        managers.append({
            "user_name": name,
            "finishes": {
                "2023": parse_finish(finish_2023),
                "2024": parse_finish(finish_2024),
                "2025": parse_finish(finish_2025)
            },
            "average_finish": round(float(avg_finish), 2) if avg_finish else None
        })

    # Sort by average finish
    managers.sort(key=lambda m: m["average_finish"] if m["average_finish"] else 999)

    return {
        "managers": managers
    }


def main():
    print(f"Loading Excel file: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Create output directory
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Extract 2024 data
    print("Extracting 2024 data...")
    data_2024 = extract_2024_data(wb)
    with open(OUTPUT_DIR / "2024.json", "w") as f:
        json.dump(data_2024, f, indent=2)
    print(f"  Saved {len(data_2024['managers'])} managers to 2024.json")

    # Extract 2023 data
    print("Extracting 2023 data...")
    data_2023 = extract_2023_data(wb)
    with open(OUTPUT_DIR / "2023.json", "w") as f:
        json.dump(data_2023, f, indent=2)
    print(f"  Saved {len(data_2023['managers'])} managers to 2023.json")

    # Extract average finishes
    print("Extracting average finishes...")
    avg_finishes = extract_average_finishes(wb)
    with open(OUTPUT_DIR / "average_finishes.json", "w") as f:
        json.dump(avg_finishes, f, indent=2)
    print(f"  Saved {len(avg_finishes['managers'])} managers to average_finishes.json")

    print("\nMigration complete!")


if __name__ == "__main__":
    main()
